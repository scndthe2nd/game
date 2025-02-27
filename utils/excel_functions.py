#excel_functions.py
"""Functions to work with Excel sheets"""

import openpyxl

def load_workbook(workbook, data_only=True, read_only=True):
    """Load an Excel Workbook"""
    return openpyxl.load_workbook(workbook, data_only=data_only, read_only=read_only)

def pull_list_from_excel(
    workbook,
    worksheet,
    header_name=None,
    row_number=None,
    column_number=None):
    """Pull List from Excel"""
    book = load_workbook(workbook)
    if not book or worksheet not in book.sheetnames:
        return []

    sheet = book[worksheet]
    cells = []

    if header_name:
        header_row_index = None
        column_index = None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == header_name:
                    header_row_index = cell.row
                    column_index = cell.column - 1
                    break
            if header_row_index is not None:
                break
        if column_index is None:
            return []
        for row in sheet.iter_rows(
            min_row=header_row_index + 1,
            min_col=column_index + 1,
            max_col=column_index + 1
            ):
            cells.append(row[0].value)
    elif row_number is not None and column_number is not None:
        for row in sheet.iter_rows(min_row=row_number + 1):
            cells.append(row[column_number].value)
    return cells

def get_value_below_header(sheet, header_name):
    """ Pull a single value from an excel sheet below a specified header """
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == header_name:
                return sheet.cell(row=cell.row + 1, column=cell.column).value
    return None

def create_mapping(keys, values, filter_keys=None):
    """Create a map between a specified key and value"""
    if filter_keys is not None:
        return {key: value for key, value in zip(keys, values) if key in filter_keys}
    return "Error, unable to map values"

# End-of-file (EOF)
